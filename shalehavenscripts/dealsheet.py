### Deal Pipeline
### Copies the deal pipeline data to the database folder, including Contact List
### and ContactBasins bridge table for Power BI.
### Developed by Michael Tanner

import os
import pandas as pd
import openpyxl


def buildDealPipeline(sourcePath):
    if not os.path.exists(sourcePath):
        raise FileNotFoundError(f"Deal sheet not found at {sourcePath}")

    destDir = os.getenv("SHALEHAVEN_DATABASE_PATH")
    os.makedirs(destDir, exist_ok=True)
    dest = os.path.join(destDir, "deal_pipeline_flat.xlsx")

    print(f"Loading {sourcePath}...")
    allSheets = pd.read_excel(sourcePath, sheet_name=None)

    # Parse Contact List from separate xlsm
    contactPath = os.getenv("SHALEHAVEN_DEAL_CONTACT")
    contactDf = None
    basinsBridge = None
    if contactPath and os.path.exists(contactPath):
        print(f"Loading contacts from {contactPath}...")
        wb = openpyxl.load_workbook(contactPath, data_only=True, keep_vba=False)
        if "Contact List" in wb.sheetnames:
            ws = wb["Contact List"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(c).strip() if c else f"Col{i}" for i, c in enumerate(rows[0])]
                contactDf = pd.DataFrame(rows[1:], columns=headers)
                contactDf = contactDf.dropna(how="all")
                contactDf.insert(0, "ContactID", range(1, len(contactDf) + 1))
                for col in contactDf.columns:
                    if any(k in col.lower() for k in ("date", "touch")):
                        contactDf[col] = pd.to_datetime(contactDf[col], errors="coerce")
                if "Basins" in contactDf.columns:
                    bridge_rows = []
                    for _, row in contactDf.iterrows():
                        basins_raw = str(row["Basins"]) if pd.notna(row["Basins"]) else ""
                        for basin in basins_raw.split(","):
                            basin = basin.strip()
                            if basin:
                                bridge_rows.append({"ContactID": row["ContactID"], "Basin": basin})
                    basinsBridge = pd.DataFrame(bridge_rows)
                    contactDf = contactDf.drop(columns=["Basins"])
                print(f"  Contact List: parsed {len(contactDf)} contacts, {len(basinsBridge) if basinsBridge is not None else 0} basin mappings")

    with pd.ExcelWriter(dest, engine="openpyxl") as writer:
        for sheetName, df in allSheets.items():
            if sheetName == "Inputs":
                continue
            df.to_excel(writer, index=False, sheet_name=sheetName)
        if contactDf is not None:
            contactDf.to_excel(writer, index=False, sheet_name="ContactList")
        if basinsBridge is not None:
            basinsBridge.to_excel(writer, index=False, sheet_name="ContactBasins")

    totalRows = sum(len(df) for df in allSheets.values())
    print(f"Wrote {len(allSheets)} sheets ({totalRows} total rows) to {dest}")
    if contactDf is not None:
        print(f"Wrote {len(contactDf)} contacts to ContactList sheet")
    if basinsBridge is not None:
        print(f"Wrote {len(basinsBridge)} basin mappings to ContactBasins sheet")
